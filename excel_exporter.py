import io
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from typing import List, Dict, Any

# Header Classification Colors (ARGB hex format)
HEADER_COLOR_MAP = {
    "NiO2": "FFD0CECE",
    "ethanol": "FFFBE5D6",
    "toluene": "FFFBE5D6",
    "ipa": "FFFBE5D6",
    "thf": "FFFBE5D6",
    "chlorobenzene": "FFFBE5D6",
    "methoxyethanol_2": "FFFBE5D6",
    "ch2cl2": "FFFBE5D6",
    "concentration": "FFE2F0D9",
    "wash": "FFDAE3F3",
    "energy_e": "FFE2F0D9",
    "cs": "FFD6DCE5",
    "fa": "FFD6DCE5",
    "ma": "FFD6DCE5",
    "pb": "FFFFF2CC",
    "sn": "FFFFF2CC",
    "i": "FFE2F0D9",
    "br": "FFE2F0D9",
    "cl": "FFE2F0D9",
    "c60": "FFD0CECE",
    "bcp": "FFD0CECE",
    "pc60bm": "FFD0CECE",
    "pcbm": "FFD0CECE",
    "pc61bm": "FFD0CECE",
    "peai": "FFD0CECE",
    "ald_sno2": "FFD0CECE",
    "pce": "FFFFCCFF",
}

# Display names for headers (35 columns)
HEADER_DISPLAY_NAMES = [
    ("ref_id", "編號"),
    ("sam_material", "SAM/HTL材料名稱"),
    ("smiles", "smile"),
    ("nio2", "NiO2"),
    ("ethanol", "ethanol"),
    ("toluene", "toluene"),
    ("ipa", "IPA"),
    ("thf", "THF"),
    ("chlorobenzene", "chlorobenzene"),
    ("methoxyethanol_2", "2-Methoxyethanol"),
    ("ch2cl2", "CH2CL2"),
    ("concentration", "concentration(mg/ml)"),
    ("wash", "wash"),
    ("energy_e", "E"),
    ("cs", "Cs"),
    ("fa", "FA"),
    ("ma", "MA"),
    ("pb", "Pb"),
    ("sn", "Sn"),
    ("i", "I"),
    ("br", "Br"),
    ("cl", "CL"),
    ("c60", "C60"),
    ("bcp", "BCP"),
    ("pc60bm", "PC60BM"),
    ("pcbm", "PCBM"),
    ("pc61bm", "PC61BM"),
    ("peai", "PEAI"),
    ("ald_sno2", "ALD-SnO2"),
    ("pce", "PCE"),
    ("reference_doi", "Reference_DOI"),
    ("ref_author", "Ref_author"),
    ("ref_journal", "Ref_journal"),
    ("data_status", "Data_status"),
    ("notes", "Notes"),
]

def generate_sam_excel(data_rows: List[Dict[str, Any]], doi_list: List[Dict[str, Any]] = None) -> bytes:
    """Generate Excel Workbook (.xlsx) with multi-tab layout and exact color codes."""
    wb = openpyxl.Workbook()
    
    # ---------------------------------------------------------
    # Tab 1: 主表 (Main Dataset)
    # ---------------------------------------------------------
    ws_main = wb.active
    ws_main.title = "主表"
    
    thin_border = Border(
        left=Side(style='thin', color='D3D3D3'),
        right=Side(style='thin', color='D3D3D3'),
        top=Side(style='thin', color='D3D3D3'),
        bottom=Side(style='thin', color='D3D3D3')
    )
    
    # 1. Write Headers (Row 1)
    for col_idx, (key, label) in enumerate(HEADER_DISPLAY_NAMES, start=1):
        cell = ws_main.cell(row=1, column=col_idx, value=label)
        cell.font = Font(name="Segoe UI", size=10, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
        
        # Header classification color
        if key in HEADER_COLOR_MAP:
            cell.fill = PatternFill(patternType="solid", fgColor=HEADER_COLOR_MAP[key])
        else:
            cell.fill = PatternFill(patternType="solid", fgColor="F2F2F2")

    # 2. Write Data Rows (Row 2+)
    for row_idx, row_data in enumerate(data_rows, start=2):
        confidence_colors = row_data.get("confidence_colors", {})
        
        for col_idx, (key, _) in enumerate(HEADER_DISPLAY_NAMES, start=1):
            val = row_data.get(key, "")
            cell = ws_main.cell(row=row_idx, column=col_idx, value=val)
            cell.font = Font(name="Segoe UI", size=10)
            cell.alignment = Alignment(vertical="center")
            cell.border = thin_border
            
            # Confidence cell color formatting
            cell_color = confidence_colors.get(key, "").lower()
            if cell_color == "red":
                cell.fill = PatternFill(patternType="solid", fgColor="FFC7CE") # Light red
            elif cell_color == "black":
                cell.fill = PatternFill(patternType="solid", fgColor="000000") # Black
                cell.font = Font(name="Segoe UI", size=10, color="FFFFFF") # White text
            elif cell_color == "white":
                cell.fill = PatternFill(patternType="solid", fgColor="FFFFFF")

    # Auto-adjust column widths
    for col in ws_main.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws_main.column_dimensions[col_letter].width = max(max_len + 3, 11)

    # ---------------------------------------------------------
    # Tab 2: 說明_Legend
    # ---------------------------------------------------------
    ws_legend = wb.create_sheet(title="說明_Legend")
    legend_data = [
        ["儲存格可信度顏色標記說明"],
        ["顏色代碼", "顏色名稱", "意義說明", "範例"],
        ["FFFFFF", "白（無填色）", "抄錄自論文/SI 文字或確定性計算", "主文明示 HOMO -5.47 eV"],
        ["FFC7CE", "紅", "需人工二次檢查：AI 讀圖取值、推算/邏輯推論、wash=0 未記載推論", "讀 Fig 1F 能階圖 / wash=0"],
        ["000000", "黑", "論文無法讀取（付費牆/僅摘要）：26 特徵全黑", "Data_status 載明原因"],
        [],
        ["標頭欄位分類底色說明"],
        ["底色", "ARGB Hex", "涵蓋欄位", "類別"],
        ["灰", "FFD0CECE", "NiO2, C60, BCP, PC60BM, PCBM, PC61BM, PEAI, ALD-SnO2", "元件堆疊材料"],
        ["橘", "FFFBE5D6", "ethanol, toluene, IPA, THF, chlorobenzene, 2-Methoxyethanol, CH2CL2", "溶劑"],
        ["綠", "FFE2F0D9", "concentration(mg/ml), E, I, Br, CL", "數值欄"],
        ["藍", "FFDAE3F3", "wash", "沖洗步驟"],
        ["藍灰", "FFD6D6E5", "Cs, FA, MA", "A-site 陽離子比"],
        ["黃", "FFFFF2CC", "Pb, Sn", "B-site 比"],
        ["粉紫", "FFFFCCFF", "PCE", "預測目標 (PCE %)"],
    ]
    for r_idx, row in enumerate(legend_data, start=1):
        for c_idx, val in enumerate(row, start=1):
            cell = ws_legend.cell(row=r_idx, column=c_idx, value=val)
            if r_idx in [1, 7]:
                cell.font = Font(name="Segoe UI", size=11, bold=True)
            else:
                cell.font = Font(name="Segoe UI", size=10)

    # ---------------------------------------------------------
    # Tab 3: 參考文獻_DOI List
    # ---------------------------------------------------------
    ws_doi = wb.create_sheet(title="參考文獻_DOI")
    ws_doi.cell(row=1, column=1, value="Index").font = Font(bold=True)
    ws_doi.cell(row=1, column=2, value="DOI").font = Font(bold=True)
    ws_doi.cell(row=1, column=3, value="URL").font = Font(bold=True)
    ws_doi.cell(row=1, column=4, value="Context / Citation").font = Font(bold=True)
    
    if doi_list:
        for idx, item in enumerate(doi_list, start=1):
            ws_doi.cell(row=idx+1, column=1, value=idx)
            ws_doi.cell(row=idx+1, column=2, value=item.get("doi", ""))
            ws_doi.cell(row=idx+1, column=3, value=item.get("url", ""))
            ws_doi.cell(row=idx+1, column=4, value=item.get("context", ""))
            
    # Auto-width for DOI tab
    for col in ws_doi.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws_doi.column_dimensions[col_letter].width = max(max_len + 3, 12)

    # ---------------------------------------------------------
    # Tab 4: 進度追蹤 (Progress tracking)
    # ---------------------------------------------------------
    ws_track = wb.create_sheet(title="進度追蹤")
    ws_track.append(["Ref編號", "DOI", "狀態", "備註"])
    ws_track.append(["1", data_rows[0].get("reference_doi", "") if data_rows else "", "Done", "已完成數據擷取與 DOI 提取"])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
