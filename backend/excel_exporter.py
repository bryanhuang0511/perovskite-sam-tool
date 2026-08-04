import io
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from typing import List, Dict, Any

HEADER_SCHEMA = [
    {"label": "Ref ID (PDF-SAM名)", "key": "ref_id", "color": "D0CECE"},
    {"label": "SAM material", "key": "sam_material", "color": "FBE5D6"},
    {"label": "SMILES", "key": "smiles", "color": "E2F0D9"},
    {"label": "NiO2", "key": "nio2", "color": "DAE3F3"},
    {"label": "Ethanol", "key": "ethanol", "color": "DAE3F3"},
    {"label": "Toluene", "key": "toluene", "color": "DAE3F3"},
    {"label": "IPA", "key": "ipa", "color": "DAE3F3"},
    {"label": "THF", "key": "thf", "color": "DAE3F3"},
    {"label": "chlorobenzene", "key": "chlorobenzene", "color": "DAE3F3"},
    {"label": "2-Methoxyethanol", "key": "methoxyethanol_2", "color": "DAE3F3"},
    {"label": "CH2CL2", "key": "ch2cl2", "color": "DAE3F3"},
    {"label": "concentration(mg/ml)", "key": "concentration", "color": "D6DCE5"},
    {"label": "wash", "key": "wash", "color": "D6DCE5"},
    {"label": "E", "key": "energy_e", "color": "FFF2CC"},
    {"label": "Cs", "key": "cs", "color": "FFCCFF"},
    {"label": "FA", "key": "fa", "color": "FFCCFF"},
    {"label": "MA", "key": "ma", "color": "FFCCFF"},
    {"label": "Pb", "key": "pb", "color": "FFCCFF"},
    {"label": "Sn", "key": "sn", "color": "FFCCFF"},
    {"label": "I", "key": "i", "color": "FFCCFF"},
    {"label": "Br", "key": "br", "color": "FFCCFF"},
    {"label": "CL", "key": "cl", "color": "FFCCFF"},
    {"label": "C60", "key": "c60", "color": "DAE3F3"},
    {"label": "BCP", "key": "bcp", "color": "DAE3F3"},
    {"label": "PC60BM", "key": "pc60bm", "color": "DAE3F3"},
    {"label": "PCBM", "key": "pcbm", "color": "DAE3F3"},
    {"label": "PC61BM", "key": "pc61bm", "color": "DAE3F3"},
    {"label": "PEAI", "key": "peai", "color": "DAE3F3"},
    {"label": "ALD-SnO2", "key": "ald_sno2", "color": "DAE3F3"},
    {"label": "PCE", "key": "pce", "color": "FFF2CC"},
    {"label": "Reference_DOI", "key": "reference_doi", "color": "D0CECE"},
    {"label": "Ref_author", "key": "ref_author", "color": "D0CECE"},
    {"label": "Ref_journal", "key": "ref_journal", "color": "D0CECE"},
    {"label": "Notes", "key": "notes", "color": "ffffff"}
]

RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
RED_FONT = Font(name="Segoe UI", size=10, color="9C0006")

BLACK_FILL = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
BLACK_FONT = Font(name="Segoe UI", size=10, color="FFFFFF")

WHITE_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
WHITE_FONT = Font(name="Segoe UI", size=10, color="000000")

THIN_BORDER = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9')
)


def generate_sam_excel(sam_dataset: List[Dict[str, Any]], doi_list: List[Dict[str, Any]] = None) -> bytes:
    """Generate Excel binary bytes containing SAM dataset sheet and DOI references sheet."""
    wb = openpyxl.Workbook()
    
    # Sheet 1: SAM Dataset
    ws_sam = wb.active
    ws_sam.title = "SAM p-i-n 特徵數據庫"
    ws_sam.views.sheetView[0].showGridLines = True

    # Write Headers
    for col_num, header in enumerate(HEADER_SCHEMA, 1):
        cell = ws_sam.cell(row=1, column=col_num, value=header["label"])
        cell.fill = PatternFill(start_color=header["color"], end_color=header["color"], fill_type="solid")
        cell.font = Font(name="Segoe UI", size=10, bold=True, color="000000")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

    ws_sam.row_dimensions[1].height = 28

    # Write SAM Rows
    for row_idx, data in enumerate(sam_dataset, start=2):
        confidence_colors = data.get("confidence_colors", {})
        if not isinstance(confidence_colors, dict):
            confidence_colors = {}

        for col_num, header in enumerate(HEADER_SCHEMA, 1):
            key = header["key"]
            val = data.get(key, "")

            if key in ["nio2", "ethanol", "toluene", "ipa", "thf", "chlorobenzene", "methoxyethanol_2",
                        "ch2cl2", "wash", "c60", "bcp", "pc60bm", "pcbm", "pc61bm", "peai", "ald_sno2"]:
                try:
                    val = int(val) if val is not None and str(val).strip() != "" else 0
                except ValueError:
                    val = 0
            elif key in ["concentration", "energy_e", "cs", "fa", "ma", "pb", "sn", "i", "br", "cl", "pce"]:
                try:
                    val = float(val) if val is not None and str(val).strip() != "" else 0.0
                except ValueError:
                    pass

            cell = ws_sam.cell(row=row_idx, column=col_num, value=val)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center" if isinstance(val, (int, float)) else "left", vertical="center")

            cell_color = confidence_colors.get(key, "white").lower()
            if cell_color == "red":
                cell.fill = RED_FILL
                cell.font = RED_FONT
            elif cell_color == "black":
                cell.fill = BLACK_FILL
                cell.font = BLACK_FONT
            else:
                cell.fill = WHITE_FILL
                cell.font = WHITE_FONT

        ws_sam.row_dimensions[row_idx].height = 22

    # Auto-adjust column widths
    for col in ws_sam.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws_sam.column_dimensions[col_letter].width = max(max_len + 3, 12)

    # Sheet 2: Reference DOIs
    if doi_list:
        ws_doi = wb.create_sheet(title="文末 References DOI 清單")
        ws_doi.views.sheetView[0].showGridLines = True

        headers_doi = ["文獻序號 #", "DOI", "完整 URL", "可信度驗證狀態", "內文引用上下文"]
        for col_num, h_text in enumerate(headers_doi, 1):
            cell = ws_doi.cell(row=1, column=col_num, value=h_text)
            cell.fill = PatternFill(start_color="38BDF8", end_color="38BDF8", fill_type="solid")
            cell.font = Font(name="Segoe UI", size=10, bold=True, color="000000")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row_idx, item in enumerate(doi_list, start=2):
            ws_doi.cell(row=row_idx, column=1, value=item.get("line_number", row_idx - 1))
            ws_doi.cell(row=row_idx, column=2, value=item.get("doi", ""))
            ws_doi.cell(row=row_idx, column=3, value=item.get("url", ""))
            ws_doi.cell(row=row_idx, column=4, value=item.get("verification", "Crossref Verifying"))
            ws_doi.cell(row=row_idx, column=5, value=item.get("context", ""))

        ws_doi.column_dimensions['A'].width = 12
        ws_doi.column_dimensions['B'].width = 30
        ws_doi.column_dimensions['C'].width = 45
        ws_doi.column_dimensions['D'].width = 35
        ws_doi.column_dimensions['E'].width = 80

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
