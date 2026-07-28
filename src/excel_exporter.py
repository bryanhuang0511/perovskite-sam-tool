import io
import re
from typing import List, Dict, Any, Union
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def generate_sam_excel(sam_data: List[Dict[str, Any]], doi_list: List[Union[str, Dict[str, Any]]], output_path: str = None) -> Union[bytes, str]:
    """
    Generate standard 35-column Excel sheet with professional header colors and confidence highlighting.
    """
    wb = openpyxl.Workbook()
    
    # --- Sheet 1: SAM Dataset ---
    ws_sam = wb.active
    ws_sam.title = "SAM Dataset"

    headers_config = [
        ("ref_id", "編號", "D0CECE"),
        ("sam_material", "SAM/HTL材料名稱", "FFFFFF"),
        ("smiles", "smile", "FFFFFF"),
        ("nio2", "NiO2", "D0CECE"),
        ("ethanol", "ethanol", "FBE5D6"),
        ("toluene", "toluene", "FBE5D6"),
        ("ipa", "IPA", "FBE5D6"),
        ("thf", "THF", "FBE5D6"),
        ("chlorobenzene", "chlorobenzene", "FBE5D6"),
        ("methoxyethanol_2", "2-Methoxyethanol", "FBE5D6"),
        ("ch2cl2", "CH2CL2", "FBE5D6"),
        ("concentration", "concentration(mg/ml)", "E2F0D9"),
        ("wash", "wash", "DAE3F3"),
        ("energy_e", "E", "E2F0D9"),
        ("cs", "Cs", "D6DCE5"),
        ("fa", "FA", "D6DCE5"),
        ("ma", "MA", "D6DCE5"),
        ("pb", "Pb", "FFF2CC"),
        ("sn", "Sn", "FFF2CC"),
        ("i", "I", "E2F0D9"),
        ("br", "Br", "E2F0D9"),
        ("cl", "CL", "E2F0D9"),
        ("c60", "C60", "D0CECE"),
        ("bcp", "BCP", "D0CECE"),
        ("pc60bm", "PC60BM", "D0CECE"),
        ("pcbm", "PCBM", "D0CECE"),
        ("pc61bm", "PC61BM", "D0CECE"),
        ("peai", "PEAI", "D0CECE"),
        ("ald_sno2", "ALD-SnO2", "D0CECE"),
        ("pce", "PCE", "FFCCFF"),
        ("reference_doi", "Reference_DOI", "FFFFFF"),
        ("ref_author", "Ref_author", "FFFFFF"),
        ("ref_journal", "Ref_journal", "FFFFFF"),
        ("data_status", "Data_status", "FFFFFF"),
        ("notes", "Notes", "FFFFFF")
    ]

    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )

    fill_red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    font_red = Font(name="Calibri", size=11, color="9C0006")
    fill_black = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
    font_black = Font(name="Calibri", size=11, color="FFFFFF")
    font_default = Font(name="Calibri", size=11, color="000000")

    for col_idx, (key, label, color_hex) in enumerate(headers_config, start=1):
        cell = ws_sam.cell(row=1, column=col_idx, value=label)
        cell.font = Font(name="Calibri", size=11, bold=True, color="000000")
        cell.fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    ws_sam.row_dimensions[1].height = 28

    for r_idx, row_data in enumerate(sam_data, start=2):
        conf_colors = row_data.get("confidence_colors", {})
        ws_sam.row_dimensions[r_idx].height = 20
        
        for c_idx, (key, label, color_hex) in enumerate(headers_config, start=1):
            val = row_data.get(key, "")
            cell = ws_sam.cell(row=r_idx, column=c_idx, value=val)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

            color_tag = (conf_colors.get(key, "")).lower()
            if color_tag == "red":
                cell.fill = fill_red
                cell.font = font_red
            elif color_tag == "black":
                cell.fill = fill_black
                cell.font = font_black
            else:
                cell.font = font_default

    for col in ws_sam.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or "")
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws_sam.column_dimensions[col_letter].width = max(max_len + 4, 12)

    # --- Sheet 2: Reference DOIs ---
    ws_doi = wb.create_sheet(title="Reference DOIs")
    ws_doi.cell(row=1, column=1, value="Index").font = Font(bold=True)
    ws_doi.cell(row=1, column=2, value="DOI (Pure Text)").font = Font(bold=True)
    ws_doi.cell(row=1, column=3, value="DOI URL Link").font = Font(bold=True)
    ws_doi.cell(row=1, column=4, value="Context Snippet").font = Font(bold=True)

    for idx, d_item in enumerate(doi_list, start=1):
        if isinstance(d_item, dict):
            doi_str = d_item.get("doi", "")
            ctx_str = d_item.get("context", "")
        else:
            doi_str = str(d_item)
            ctx_str = ""

        ws_doi.cell(row=idx+1, column=1, value=idx)
        cell_doi = ws_doi.cell(row=idx+1, column=2, value=doi_str)
        cell_doi.number_format = '@'  # Explicit text format
        
        cell_link = ws_doi.cell(row=idx+1, column=3, value=f"https://doi.org/{doi_str}")
        cell_link.hyperlink = f"https://doi.org/{doi_str}"
        cell_link.font = Font(color="0563C1", underline="single")

        ws_doi.cell(row=idx+1, column=4, value=ctx_str)

    ws_doi.column_dimensions['A'].width = 8
    ws_doi.column_dimensions['B'].width = 35
    ws_doi.column_dimensions['C'].width = 45
    ws_doi.column_dimensions['D'].width = 60

    if output_path:
        wb.save(output_path)
        return output_path
    else:
        output_stream = io.BytesIO()
        wb.save(output_stream)
        return output_stream.getvalue()
